"""Build one Excel workbook per project: raw data + KPI dashboard + analysis tabs.
Calculations use Excel formulas (SUMIFS/COUNTIFS/AVERAGEIFS) so workbooks stay live."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

BASE = "/sessions/stoic-beautiful-archimedes/mnt/Portfolio"
FONT = "Arial"
NAVY = "1F3864"; BLUE = "2E5496"; LIGHT = "D9E1F2"; GREY = "F2F2F2"
WHITE = "FFFFFF"

def style_header(ws, row, ncols, fill=NAVY):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def title(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1,1,text)
    c.font = Font(name=FONT, bold=True, size=16, color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

def write_raw(wb, df, name):
    ws = wb.create_sheet(name)
    ws.append(list(df.columns))
    style_header(ws, 1, len(df.columns), fill=BLUE)
    for r in df.itertuples(index=False):
        ws.append(list(r))
    ws.freeze_panes = "A2"
    for i, col in enumerate(df.columns, 1):
        w = min(max(len(str(col))+2, 12), 22)
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

def kpi_card(ws, row, col, label, formula, numfmt):
    lab = ws.cell(row, col, label)
    lab.font = Font(name=FONT, bold=True, size=10, color="595959")
    lab.fill = PatternFill("solid", fgColor=GREY)
    lab.alignment = Alignment(horizontal="center")
    val = ws.cell(row+1, col, formula)
    val.font = Font(name=FONT, bold=True, size=18, color=NAVY)
    val.fill = PatternFill("solid", fgColor=LIGHT)
    val.alignment = Alignment(horizontal="center")
    val.number_format = numfmt
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)

# ============================================================
# PROJECT 1 — READMISSIONS
# ============================================================
def build_readmissions():
    pat = pd.read_csv(f"{BASE}/01_hospital_readmissions/data/patients.csv")
    adm = pd.read_csv(f"{BASE}/01_hospital_readmissions/data/admissions.csv")
    wb = Workbook(); wb.remove(wb.active)

    # Dashboard
    dash = wb.create_sheet("Dashboard")
    title(dash, "Hospital Readmissions — Executive Dashboard", 8)
    n = len(adm); R = "Admissions"          # raw sheet name
    rng_read = f"{R}!J2:J{n+1}"             # readmitted_30d col J
    rng_disp = f"{R}!H2:H{n+1}"             # discharge_disposition col H
    rng_los  = f"{R}!E2:E{n+1}"             # length_of_stay col E
    rng_chg  = f"{R}!I2:I{n+1}"             # total_charges col I
    kpi_card(dash,3,1,"Total Discharges", f'=COUNTIF({rng_disp},"<>Expired")', "#,##0")
    kpi_card(dash,3,3,"30-Day Readmissions", f"=SUM({rng_read})", "#,##0")
    kpi_card(dash,3,5,"Readmission Rate", f'=SUM({rng_read})/COUNTIF({rng_disp},"<>Expired")', "0.0%")
    kpi_card(dash,3,7,"Avg Length of Stay", f"=AVERAGE({rng_los})", "0.0")
    kpi_card(dash,6,1,"Avg Charge / Admission", f"=AVERAGE({rng_chg})", '$#,##0')
    kpi_card(dash,6,3,"Est. Readmission Charges", f"=SUM({rng_read})*AVERAGE({rng_chg})", '$#,##0')
    kpi_card(dash,6,5,"Unique Patients", f"={len(pat)}", "#,##0")
    kpi_card(dash,6,7,"Avg Patient Age", f"=AVERAGE(Patients!D2:D{len(pat)+1})", "0.0")
    for col in range(1,9): dash.column_dimensions[get_column_letter(col)].width = 15

    # By diagnosis sheet
    diags = sorted(adm.primary_diagnosis.unique())
    bydx = wb.create_sheet("By_Diagnosis")
    title(bydx, "Readmission Rate by Primary Diagnosis", 5)
    bydx.append([]); hdr_row = 3
    for i,h in enumerate(["Primary Diagnosis","Discharges","Readmissions","Readmission Rate","Avg LOS"],1):
        bydx.cell(hdr_row,i,h)
    style_header(bydx, hdr_row, 5)
    rng_dx = f"{R}!F2:F{n+1}"
    for k,d in enumerate(diags):
        r = hdr_row+1+k
        bydx.cell(r,1,d)
        bydx.cell(r,2, f'=COUNTIFS({rng_dx},A{r},{rng_disp},"<>Expired")')
        bydx.cell(r,3, f'=SUMIFS({rng_read},{rng_dx},A{r},{rng_disp},"<>Expired")')
        bydx.cell(r,4, f"=IFERROR(C{r}/B{r},0)"); bydx.cell(r,4).number_format="0.0%"
        bydx.cell(r,5, f"=AVERAGEIFS({rng_los},{rng_dx},A{r})"); bydx.cell(r,5).number_format="0.0"
    last = hdr_row+len(diags)
    for col in range(1,6): bydx.column_dimensions[get_column_letter(col)].width = 20
    ch = BarChart(); ch.type="bar"; ch.title="Readmission Rate by Diagnosis"; ch.height=9; ch.width=18
    data = Reference(bydx, min_col=4, min_row=hdr_row, max_row=last)
    cats = Reference(bydx, min_col=1, min_row=hdr_row+1, max_row=last)
    ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
    bydx.add_chart(ch, "G3")

    # By LOS band sheet
    bylos = wb.create_sheet("By_LOS_Band")
    title(bylos, "Readmission Rate by Length of Stay", 4)
    hdr_row=3
    for i,h in enumerate(["LOS Band","Discharges","Readmissions","Readmission Rate"],1):
        bylos.cell(hdr_row,i,h)
    style_header(bylos,hdr_row,4)
    bands = [("1-2 days",1,2),("3-5 days",3,5),("6-9 days",6,9),("10+ days",10,999)]
    for k,(lbl,lo,hi) in enumerate(bands):
        r=hdr_row+1+k
        bylos.cell(r,1,lbl)
        bylos.cell(r,2, f'=COUNTIFS({rng_los},">="&{lo},{rng_los},"<="&{hi},{rng_disp},"<>Expired")')
        bylos.cell(r,3, f'=SUMIFS({rng_read},{rng_los},">="&{lo},{rng_los},"<="&{hi},{rng_disp},"<>Expired")')
        bylos.cell(r,4, f"=IFERROR(C{r}/B{r},0)"); bylos.cell(r,4).number_format="0.0%"
    for col in range(1,5): bylos.column_dimensions[get_column_letter(col)].width=18
    ch2=BarChart(); ch2.title="Readmission Rate by LOS Band"; ch2.height=8; ch2.width=15
    d2=Reference(bylos,min_col=4,min_row=hdr_row,max_row=hdr_row+len(bands))
    c2=Reference(bylos,min_col=1,min_row=hdr_row+1,max_row=hdr_row+len(bands))
    ch2.add_data(d2,titles_from_data=True); ch2.set_categories(c2); bylos.add_chart(ch2,"F3")

    write_raw(wb, adm, "Admissions")
    write_raw(wb, pat, "Patients")
    wb.move_sheet("Dashboard", -(len(wb.sheetnames)-1))
    wb.save(f"{BASE}/01_hospital_readmissions/excel/readmissions_analysis.xlsx")

# ============================================================
# PROJECT 2 — ED THROUGHPUT
# ============================================================
def build_ed():
    ed = pd.read_csv(f"{BASE}/02_ed_throughput/data/ed_visits.csv")
    wb = Workbook(); wb.remove(wb.active)
    n=len(ed); R="ED_Visits"
    # columns: A visit_id B arrival_dt C arrival_hour D dow E mode F esi G chief H age I gender
    # J door_to_triage K triage_to_provider L ed_los M disposition N lwbs O boarding
    c_hour=f"{R}!C2:C{n+1}"; c_esi=f"{R}!F2:F{n+1}"; c_d2t=f"{R}!J2:J{n+1}"
    c_t2p=f"{R}!K2:K{n+1}"; c_los=f"{R}!L2:L{n+1}"; c_disp=f"{R}!M2:M{n+1}"
    c_lwbs=f"{R}!N2:N{n+1}"; c_board=f"{R}!O2:O{n+1}"; c_dow=f"{R}!D2:D{n+1}"

    dash=wb.create_sheet("Dashboard")
    title(dash,"Emergency Department Throughput — Dashboard",8)
    kpi_card(dash,3,1,"Total Visits", f"=COUNT({R}!A2:A{n+1})","#,##0")
    kpi_card(dash,3,3,"Avg Door-to-Triage (min)", f"=AVERAGE({c_d2t})","0.0")
    kpi_card(dash,3,5,"Avg Wait to Provider (min)", f"=AVERAGE({c_t2p})","0.0")
    kpi_card(dash,3,7,"Avg ED LOS (min)", f"=AVERAGE({c_los})","0.0")
    kpi_card(dash,6,1,"LWBS Rate", f"=SUM({c_lwbs})/COUNT({R}!A2:A{n+1})","0.0%")
    kpi_card(dash,6,3,"Admit Rate", f'=COUNTIF({c_disp},"Admitted")/COUNT({R}!A2:A{n+1})',"0.0%")
    kpi_card(dash,6,5,"Avg Boarding (min)", f'=AVERAGEIF({c_disp},"Admitted",{c_board})',"0.0")
    kpi_card(dash,6,7,"% LOS Under 4h", f'=COUNTIF({c_los},"<=240")/COUNT({R}!A2:A{n+1})',"0.0%")
    for col in range(1,9): dash.column_dimensions[get_column_letter(col)].width=15

    # By ESI
    bye=wb.create_sheet("By_Triage_ESI")
    title(bye,"Wait & Volume by ESI Triage Level",4)
    hdr=3
    for i,h in enumerate(["ESI Level","Visits","Avg Wait to Provider (min)","Avg ED LOS (min)"],1):
        bye.cell(hdr,i,h)
    style_header(bye,hdr,4)
    for k,lvl in enumerate([1,2,3,4,5]):
        r=hdr+1+k
        bye.cell(r,1,lvl)
        bye.cell(r,2, f"=COUNTIF({c_esi},A{r})")
        bye.cell(r,3, f"=AVERAGEIFS({c_t2p},{c_esi},A{r})"); bye.cell(r,3).number_format="0.0"
        bye.cell(r,4, f"=AVERAGEIFS({c_los},{c_esi},A{r})"); bye.cell(r,4).number_format="0.0"
    for col in range(1,5): bye.column_dimensions[get_column_letter(col)].width=22
    ch=BarChart(); ch.title="Avg Wait to Provider by ESI Level"; ch.height=8; ch.width=15
    d=Reference(bye,min_col=3,min_row=hdr,max_row=hdr+5); c=Reference(bye,min_col=1,min_row=hdr+1,max_row=hdr+5)
    ch.add_data(d,titles_from_data=True); ch.set_categories(c); bye.add_chart(ch,"F3")

    # By hour
    byh=wb.create_sheet("By_Hour")
    title(byh,"Volume & Wait by Arrival Hour",4)
    hdr=3
    for i,h in enumerate(["Hour","Visits","Avg Wait (min)","Avg LOS (min)"],1):
        byh.cell(hdr,i,h)
    style_header(byh,hdr,4)
    for hh in range(24):
        r=hdr+1+hh
        byh.cell(r,1,hh)
        byh.cell(r,2, f"=COUNTIF({c_hour},A{r})")
        byh.cell(r,3, f"=IFERROR(AVERAGEIFS({c_t2p},{c_hour},A{r}),0)"); byh.cell(r,3).number_format="0.0"
        byh.cell(r,4, f"=IFERROR(AVERAGEIFS({c_los},{c_hour},A{r}),0)"); byh.cell(r,4).number_format="0.0"
    for col in range(1,5): byh.column_dimensions[get_column_letter(col)].width=16
    lc=LineChart(); lc.title="ED Visits by Hour of Day"; lc.height=9; lc.width=18
    d=Reference(byh,min_col=2,min_row=hdr,max_row=hdr+24); c=Reference(byh,min_col=1,min_row=hdr+1,max_row=hdr+24)
    lc.add_data(d,titles_from_data=True); lc.set_categories(c); byh.add_chart(lc,"F3")

    # By day of week
    byd=wb.create_sheet("By_DayOfWeek")
    title(byd,"Volume & LWBS by Day of Week",4)
    hdr=3
    for i,h in enumerate(["Day","Visits","Avg LOS (min)","LWBS Rate"],1):
        byd.cell(hdr,i,h)
    style_header(byd,hdr,4)
    for k,day in enumerate(["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]):
        r=hdr+1+k
        byd.cell(r,1,day)
        byd.cell(r,2, f'=COUNTIF({c_dow},A{r})')
        byd.cell(r,3, f'=AVERAGEIFS({c_los},{c_dow},A{r})'); byd.cell(r,3).number_format="0.0"
        byd.cell(r,4, f'=SUMIFS({c_lwbs},{c_dow},A{r})/COUNTIF({c_dow},A{r})'); byd.cell(r,4).number_format="0.0%"
    for col in range(1,5): byd.column_dimensions[get_column_letter(col)].width=16

    write_raw(wb, ed, "ED_Visits")
    wb.move_sheet("Dashboard", -(len(wb.sheetnames)-1))
    wb.save(f"{BASE}/02_ed_throughput/excel/ed_throughput_analysis.xlsx")

# ============================================================
# PROJECT 3 — CLAIMS DENIALS
# ============================================================
def build_claims():
    cl = pd.read_csv(f"{BASE}/03_claims_denials/data/claims.csv")
    wb = Workbook(); wb.remove(wb.active)
    n=len(cl); R="Claims"
    # A claim_id B service_date C submit_date D paid_date E payer F department G cpt
    # H billed I paid J status K is_denied L denial_reason M days_to_payment
    c_payer=f"{R}!E2:E{n+1}"; c_dept=f"{R}!F2:F{n+1}"; c_billed=f"{R}!H2:H{n+1}"
    c_paid=f"{R}!I2:I{n+1}"; c_status=f"{R}!J2:J{n+1}"; c_den=f"{R}!K2:K{n+1}"
    c_reason=f"{R}!L2:L{n+1}"; c_days=f"{R}!M2:M{n+1}"

    dash=wb.create_sheet("Dashboard")
    title(dash,"Claims Denials & Revenue Cycle — Dashboard",8)
    kpi_card(dash,3,1,"Total Claims", f"=COUNT({R}!A2:A{n+1})","#,##0")
    kpi_card(dash,3,3,"Total Billed", f"=SUM({c_billed})",'$#,##0')
    kpi_card(dash,3,5,"Total Collected", f"=SUM({c_paid})",'$#,##0')
    kpi_card(dash,3,7,"Net Collection Rate", f"=SUM({c_paid})/SUM({c_billed})","0.0%")
    kpi_card(dash,6,1,"Denial Rate", f"=SUM({c_den})/COUNT({R}!A2:A{n+1})","0.0%")
    kpi_card(dash,6,3,"Denied Claims", f"=SUM({c_den})","#,##0")
    kpi_card(dash,6,5,"Revenue Gap (Billed-Paid)", f"=SUM({c_billed})-SUM({c_paid})",'$#,##0')
    kpi_card(dash,6,7,"Avg Days to Payment", f'=AVERAGEIF({c_days},">0")',"0.0")
    for col in range(1,9): dash.column_dimensions[get_column_letter(col)].width=15

    # By payer
    payers=sorted(cl.payer.unique())
    byp=wb.create_sheet("By_Payer")
    title(byp,"Denial Rate & Collections by Payer",5)
    hdr=3
    for i,h in enumerate(["Payer","Claims","Denial Rate","Billed","Net Collection Rate"],1):
        byp.cell(hdr,i,h)
    style_header(byp,hdr,5)
    for k,p in enumerate(payers):
        r=hdr+1+k
        byp.cell(r,1,p)
        byp.cell(r,2, f"=COUNTIF({c_payer},A{r})")
        byp.cell(r,3, f"=SUMIFS({c_den},{c_payer},A{r})/COUNTIF({c_payer},A{r})"); byp.cell(r,3).number_format="0.0%"
        byp.cell(r,4, f"=SUMIFS({c_billed},{c_payer},A{r})"); byp.cell(r,4).number_format='$#,##0'
        byp.cell(r,5, f"=SUMIFS({c_paid},{c_payer},A{r})/SUMIFS({c_billed},{c_payer},A{r})"); byp.cell(r,5).number_format="0.0%"
    last=hdr+len(payers)
    for col in range(1,6): byp.column_dimensions[get_column_letter(col)].width=20
    ch=BarChart(); ch.title="Denial Rate by Payer"; ch.height=9; ch.width=16
    d=Reference(byp,min_col=3,min_row=hdr,max_row=last); c=Reference(byp,min_col=1,min_row=hdr+1,max_row=last)
    ch.add_data(d,titles_from_data=True); ch.set_categories(c); byp.add_chart(ch,"G3")

    # By denial reason
    reasons=sorted([x for x in cl.denial_reason.unique() if isinstance(x,str) and x])
    byr=wb.create_sheet("By_Denial_Reason")
    title(byr,"Denials by Reason (Pareto)",4)
    hdr=3
    for i,h in enumerate(["Denial Reason","Denied Claims","Billed at Risk","% of Denials"],1):
        byr.cell(hdr,i,h)
    style_header(byr,hdr,4)
    for k,rs in enumerate(reasons):
        r=hdr+1+k
        byr.cell(r,1,rs)
        byr.cell(r,2, f"=COUNTIFS({c_reason},A{r})")
        byr.cell(r,3, f"=SUMIFS({c_billed},{c_reason},A{r})"); byr.cell(r,3).number_format='$#,##0'
        byr.cell(r,4, f"=B{r}/SUM({c_den})"); byr.cell(r,4).number_format="0.0%"
    last=hdr+len(reasons)
    for col in range(1,5): byr.column_dimensions[get_column_letter(col)].width=22
    ch2=BarChart(); ch2.type="bar"; ch2.title="Denied Claims by Reason"; ch2.height=9; ch2.width=16
    d=Reference(byr,min_col=2,min_row=hdr,max_row=last); c=Reference(byr,min_col=1,min_row=hdr+1,max_row=last)
    ch2.add_data(d,titles_from_data=True); ch2.set_categories(c); byr.add_chart(ch2,"F3")

    # By department
    depts=sorted(cl.department.unique())
    byd=wb.create_sheet("By_Department")
    title(byd,"Denial Rate & Revenue Gap by Department",4)
    hdr=3
    for i,h in enumerate(["Department","Claims","Denial Rate","Revenue Gap"],1):
        byd.cell(hdr,i,h)
    style_header(byd,hdr,4)
    for k,dp in enumerate(depts):
        r=hdr+1+k
        byd.cell(r,1,dp)
        byd.cell(r,2, f"=COUNTIF({c_dept},A{r})")
        byd.cell(r,3, f"=SUMIFS({c_den},{c_dept},A{r})/COUNTIF({c_dept},A{r})"); byd.cell(r,3).number_format="0.0%"
        byd.cell(r,4, f"=SUMIFS({c_billed},{c_dept},A{r})-SUMIFS({c_paid},{c_dept},A{r})"); byd.cell(r,4).number_format='$#,##0'
    for col in range(1,5): byd.column_dimensions[get_column_letter(col)].width=20

    write_raw(wb, cl, "Claims")
    wb.move_sheet("Dashboard", -(len(wb.sheetnames)-1))
    wb.save(f"{BASE}/03_claims_denials/excel/claims_denials_analysis.xlsx")

build_readmissions(); print("readmissions xlsx ok")
build_ed(); print("ed xlsx ok")
build_claims(); print("claims xlsx ok")
